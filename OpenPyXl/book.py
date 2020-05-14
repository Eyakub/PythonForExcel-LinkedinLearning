from openpyxl.workbook import Workbook
from openpyxl import load_workbook

wb = Workbook() # workbook = wb
ws = wb.active # ws = worksheet

ws1 = wb.create_sheet('NewSheet')
ws2 = wb.create_sheet('Another', 0)

ws.title = 'MySheet'

wb2 = load_workbook('regions.xlsx')

new_sheet = wb2.create_sheet('NewSheet')
active_sheet = wb2.active

active_sheet['A1'] = 0
wb2.save('modified.xlsx')